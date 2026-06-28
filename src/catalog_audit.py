"""
catalog_audit.py
----------------
Read-only catalog **standardization audit**. Walks the live Shopee + TikTok Shop
catalogs and flags every base SKU that does not yet meet the standardization
rules, writing the result to an Excel workbook (a report — it changes nothing).

Rules
-----
1. Shopee: each listing should have at least `GROSIR_MIN_LAYERS` (3) Harga Grosir
   tiers. Readable via `get_item_base_info.wholesales`; NOT writable via the Open
   API (set manually in Seller Center).
2. TikTok Shop: a SKU whose base (1PCS) price <= `TIKTOK_LOW_PRICE_THRESHOLD`
   (Rp5.000) should have at least one pack-size variant (xxPCS) beyond 1PCS.
   Fixable via `/variant_set`.
3. Shopee minimum purchase = `ceil(SHOPEE_MIN_BUY_IDR / base_price)` units. The
   current min-buy isn't exposed by the Open API, so the audit reports the
   TARGET to set manually.
4. TikTok Shop minimum purchase = `TIKTOK_MIN_BUY_LOW` (2) when base price
   <= Rp5.000 (else 1). Current min-buy isn't read today, so the audit reports
   the TARGET.

Only rules 1 & 2 are verifiable today, so the "Tidak Standar" sheet lists SKUs
failing those; the min-buy targets (rules 3 & 4) are a separate worklist sheet.
"""

from __future__ import annotations

import math
import time
from datetime import datetime, timedelta, timezone

from src import config, shopee_client, telegram_sender, tiktokshop_client
from src.shopee_detail_enrichment import _extract_price_idr
from src.variant_set_tiktok import BUBBLE_WRAP_SELLER_SKU, _sku_price_idr

# Tunables (standardization targets).
GROSIR_MIN_LAYERS = 3
SHOPEE_MIN_BUY_IDR = 20000
TIKTOK_LOW_PRICE_THRESHOLD = 5000
TIKTOK_MIN_BUY_LOW = 2

_SHOPEE_BASE_INFO_BATCH = 50
_WIB = timezone(timedelta(hours=7))


# ----------------------------------------------------------------------
# Pure rule logic
# ----------------------------------------------------------------------
def shopee_min_buy_units(base_price: int | None, idr: int = SHOPEE_MIN_BUY_IDR) -> int | None:
    """Target Shopee minimum purchase: `ceil(idr / base_price)` units."""
    if not base_price or base_price <= 0:
        return None
    return math.ceil(idr / base_price)


def tiktok_min_buy_target(
        base_price: int | None,
        threshold: int = TIKTOK_LOW_PRICE_THRESHOLD,
        low: int = TIKTOK_MIN_BUY_LOW,
) -> int | None:
    """Target TikTok min purchase: `low` when base price <= threshold, else 1."""
    if base_price is None:
        return None
    return low if base_price <= threshold else 1


def needs_tiktok_packs(base_price: int | None, threshold: int = TIKTOK_LOW_PRICE_THRESHOLD) -> bool:
    """Rule 2 applies only to TikTok SKUs whose base price is <= threshold."""
    return base_price is not None and base_price <= threshold


def has_pack_variant(variants: list[dict]) -> bool:
    """True if any non-Bubble-Wrap variant has multiplier > 1 (a real pack)."""
    for v in variants:
        if (v.get("raw_sku") or "").upper() == BUBBLE_WRAP_SELLER_SKU:
            continue
        if int(v.get("multiplier") or 1) > 1:
            return True
    return False


def grosir_ok(num_layers: int | None, minimum: int = GROSIR_MIN_LAYERS) -> bool:
    """Rule 1: the Shopee listing has at least `minimum` Harga Grosir tiers."""
    return (num_layers or 0) >= minimum


def audit_sku(
        base_sku: str,
        shopee_variants: list[dict],
        tiktok_variants: list[dict],
        shopee_price: int | None,
        grosir_layers: int | None,
        tiktok_price: int | None,
) -> dict:
    """Evaluate one base SKU against every applicable rule. Pure."""
    on_shopee = bool(shopee_variants)
    on_tiktok = bool(tiktok_variants)
    violations: list[str] = []

    rule1: bool | None = None
    if on_shopee:
        rule1 = grosir_ok(grosir_layers)
        if not rule1:
            violations.append(
                f"Shopee Harga Grosir {grosir_layers or 0} layer (min {GROSIR_MIN_LAYERS})"
            )

    rule2: bool | None = None
    if on_tiktok and needs_tiktok_packs(tiktok_price):
        rule2 = has_pack_variant(tiktok_variants)
        if not rule2:
            violations.append("TikTok Shop belum punya varian pack (xxPCS)")

    return {
        "base_sku": base_sku,
        "on_shopee": on_shopee,
        "on_tiktok": on_tiktok,
        "shopee_price": shopee_price,
        "grosir_layers": grosir_layers if on_shopee else None,
        "rule1_grosir_ok": rule1,
        "tiktok_price": tiktok_price,
        "tiktok_pack_ok": rule2,
        "shopee_min_buy_target": shopee_min_buy_units(shopee_price) if on_shopee else None,
        "tiktok_min_buy_target": tiktok_min_buy_target(tiktok_price) if on_tiktok else None,
        "violations": violations,
    }


# ----------------------------------------------------------------------
# Data gathering (live, read-only)
# ----------------------------------------------------------------------
def _one_pcs_variant(variants: list[dict]) -> dict | None:
    """The 1PCS (multiplier==1) variant, else the smallest-multiplier one."""
    if not variants:
        return None
    ones = [v for v in variants if int(v.get("multiplier") or 1) == 1]
    return ones[0] if ones else variants[0]


def _gather_shopee_info(shopee_catalog: dict[str, list[dict]]) -> dict[str, dict]:
    """base_sku -> {price, grosir_layers} from one batched get_item_base_info pass."""
    base_to_item: dict[str, int] = {}
    for base_sku, variants in shopee_catalog.items():
        one = _one_pcs_variant(variants)
        if one and one.get("item_id"):
            base_to_item[base_sku] = int(one["item_id"])

    item_ids = sorted(set(base_to_item.values()))
    by_item: dict[int, dict] = {}
    for start in range(0, len(item_ids), _SHOPEE_BASE_INFO_BATCH):
        batch = item_ids[start:start + _SHOPEE_BASE_INFO_BATCH]
        data = shopee_client._signed_get(  # noqa: SLF001 - internal client reuse
            "/api/v2/product/get_item_base_info",
            {"item_id_list": ",".join(str(i) for i in batch)},
        )
        for item in (data.get("response") or {}).get("item_list", []):
            by_item[int(item["item_id"])] = {
                "price": _extract_price_idr(item),
                "grosir_layers": len(item.get("wholesales") or []),
            }
        time.sleep(config.DELAY_BETWEEN_CALLS_SECONDS)

    return {b: by_item.get(iid, {}) for b, iid in base_to_item.items()}


def _gather_tiktok_prices(tiktok_catalog: dict[str, list[dict]]) -> dict[str, int | None]:
    """base_sku -> 1PCS price, via one product-detail read per unique product."""
    base_to_sku: dict[str, tuple] = {}
    product_ids: set = set()
    for base_sku, variants in tiktok_catalog.items():
        one = _one_pcs_variant(variants)
        if one:
            base_to_sku[base_sku] = (one.get("product_id"), one.get("sku_id"))
            if one.get("product_id"):
                product_ids.add(one["product_id"])

    price_by_sku: dict = {}
    for pid in product_ids:
        try:
            detail = tiktokshop_client.fetch_product_detail_raw(pid)
        except Exception as e:  # noqa: BLE001 - audit is best-effort per product
            print(f"  [audit] TikTok detail {pid} failed: {e}")
            continue
        for s in (detail.get("skus") or []):
            price_by_sku[s.get("id")] = _sku_price_idr(s)
        time.sleep(config.DELAY_BETWEEN_CALLS_SECONDS)

    return {b: price_by_sku.get(sid) for b, (pid, sid) in base_to_sku.items()}


# ----------------------------------------------------------------------
# Runner
# ----------------------------------------------------------------------
def run_catalog_audit(output_path: str) -> int:
    print("=" * 70)
    print("ITBisa Shop Stock Bot — Catalog Standardization Audit (read-only)")
    print("=" * 70)

    try:
        print("Walking Shopee catalog...")
        shopee_catalog = shopee_client.fetch_catalog()
        print("Walking TikTok Shop catalog...")
        tiktok_catalog = tiktokshop_client.fetch_catalog()
        shopee_info = _gather_shopee_info(shopee_catalog)
        tiktok_prices = _gather_tiktok_prices(tiktok_catalog)
    except Exception as e:  # noqa: BLE001 - surface any walk/auth failure to Telegram
        msg = f"Gagal membaca katalog untuk audit: {e}"
        print(f"✗ {msg}")
        telegram_sender.send_alert(msg, mode="Audit")
        return 1

    rows = []
    for base_sku in sorted(set(shopee_catalog) | set(tiktok_catalog)):
        info = shopee_info.get(base_sku) or {}
        rows.append(audit_sku(
            base_sku,
            shopee_catalog.get(base_sku) or [],
            tiktok_catalog.get(base_sku) or [],
            shopee_price=info.get("price"),
            grosir_layers=info.get("grosir_layers", 0),
            tiktok_price=tiktok_prices.get(base_sku),
        ))

    _write_audit_workbook(output_path, rows)

    not_standard = [r for r in rows if r["violations"]]
    rule1_fail = sum(1 for r in rows if r["rule1_grosir_ok"] is False)
    rule2_fail = sum(1 for r in rows if r["tiktok_pack_ok"] is False)
    print(f"Total SKU: {len(rows)} | tidak standar (rule 1/2): {len(not_standard)} "
          f"| Shopee grosir<3: {rule1_fail} | TikTok tanpa pack: {rule2_fail}")
    print(f"✓ Excel: {output_path}")

    telegram_sender.send_catalog_audit_summary({
        "total": len(rows),
        "not_standard": len(not_standard),
        "rule1_fail": rule1_fail,
        "rule2_fail": rule2_fail,
        "output_name": output_path,
    })
    return 0


# ----------------------------------------------------------------------
# Excel output
# ----------------------------------------------------------------------
def _fmt_price(value: int | None) -> str:
    if value is None:
        return "—"
    return f"Rp{value:,}".replace(",", ".")


def _write_audit_workbook(path: str, rows: list[dict]) -> None:
    # Imported lazily so the pure rule logic can be unit-tested without openpyxl.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    title_font = Font(bold=True, size=14)
    generated = datetime.now(_WIB).strftime("%Y-%m-%d %H:%M WIB")

    def _header(ws, headers, widths):
        ws.append(headers)            # header as its own row (row 3: after title + blank)
        row = ws.max_row
        for i, w in enumerate(widths, start=1):
            c = ws.cell(row=row, column=i)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[chr(64 + i)].width = w

    wb = Workbook()

    # ---- 00_Summary ----
    ws = wb.active
    ws.title = "00_Summary"
    ws["A1"] = "AUDIT STANDARDISASI KATALOG — ITBISA SHOP"
    ws["A1"].font = title_font
    ws["A2"] = f"Dibuat: {generated} (laporan saja — tidak mengubah apa pun)"
    not_standard = [r for r in rows if r["violations"]]
    summary = [
        ("Total SKU", len(rows)),
        ("Tidak standar (rule 1/2)", len(not_standard)),
        ("Shopee Harga Grosir < 3 layer (rule 1)", sum(1 for r in rows if r["rule1_grosir_ok"] is False)),
        ("TikTok ≤Rp5.000 tanpa varian pack (rule 2)", sum(1 for r in rows if r["tiktok_pack_ok"] is False)),
        ("Di Shopee", sum(1 for r in rows if r["on_shopee"])),
        ("Di TikTok Shop", sum(1 for r in rows if r["on_tiktok"])),
    ]
    for i, (label, value) in enumerate(summary, start=4):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value).font = Font(bold=True)
    ws.cell(row=11, column=1, value="Catatan:")
    notes = [
        "Rule 1 (Harga Grosir) & rule 3 (min beli Shopee) hanya bisa di-set MANUAL di Seller Center "
        "(Open API Shopee tidak mendukung).",
        "Rule 2 (varian pack TikTok) bisa diperbaiki via /variant_set.",
        "Rule 3 & 4 (min beli): nilai sekarang tidak terbaca via API — kolom adalah TARGET yang harus di-set.",
    ]
    for i, n in enumerate(notes, start=12):
        ws.cell(row=i, column=1, value=f"• {n}")
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 14

    # ---- 01_Tidak_Standar (rules 1 & 2 — verifiable failures) ----
    ws = wb.create_sheet("01_Tidak_Standar")
    ws.append([f"SKU tidak standar (rule 1 / rule 2) — {generated}"])
    ws["A1"].font = title_font
    ws.append([])
    headers = ["Base SKU", "Shopee?", "TikTok?", "Harga Shopee", "Grosir layer",
               "Rule 1 OK?", "Harga TikTok", "Ada pack?", "Rule 2 OK?", "Pelanggaran"]
    widths = [40, 9, 9, 14, 12, 11, 14, 10, 11, 50]
    _header(ws, headers, widths)
    for r in not_standard:
        ws.append([
            r["base_sku"],
            "Ya" if r["on_shopee"] else "—",
            "Ya" if r["on_tiktok"] else "—",
            _fmt_price(r["shopee_price"]),
            r["grosir_layers"] if r["grosir_layers"] is not None else "—",
            _rule_cell(r["rule1_grosir_ok"]),
            _fmt_price(r["tiktok_price"]),
            _pack_cell(r["tiktok_pack_ok"]),
            _rule_cell(r["tiktok_pack_ok"]),
            "; ".join(r["violations"]),
        ])
    ws.freeze_panes = "A4"

    # ---- 02_Target_Min_Buy (rules 3 & 4 — set manually) ----
    ws = wb.create_sheet("02_Target_Min_Buy")
    ws.append([f"Target minimum pembelian — set manual ({generated})"])
    ws["A1"].font = title_font
    ws.append([])
    headers = ["Base SKU", "Harga Shopee", "Target Min Beli Shopee (unit)",
               "Harga TikTok", "Target Min Beli TikTok (unit)"]
    widths = [40, 14, 28, 14, 28]
    _header(ws, headers, widths)
    for r in rows:
        if not (r["on_shopee"] or r["on_tiktok"]):
            continue
        ws.append([
            r["base_sku"],
            _fmt_price(r["shopee_price"]) if r["on_shopee"] else "—",
            r["shopee_min_buy_target"] if r["shopee_min_buy_target"] is not None else "—",
            _fmt_price(r["tiktok_price"]) if r["on_tiktok"] else "—",
            r["tiktok_min_buy_target"] if r["tiktok_min_buy_target"] is not None else "—",
        ])
    ws.freeze_panes = "A4"

    wb.save(path)


def _rule_cell(ok: bool | None) -> str:
    if ok is None:
        return "n/a"
    return "OK" if ok else "❌"


def _pack_cell(ok: bool | None) -> str:
    if ok is None:
        return "n/a"
    return "Ya" if ok else "Tidak"
